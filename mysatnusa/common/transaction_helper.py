from django.db import connection, transaction
from django.conf import settings
import os
from datetime import datetime,timedelta
from datetime import datetime, date 
import re
import json
import traceback
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from urllib.parse import urlparse, parse_qs

# Helper: Execute query 
def execute_query(sql_query, params=None):
    """
    Executes a raw SQL query and returns the results as a list of dictionaries.

    Args:
        sql_query (str): The SQL query to execute.
        params (tuple or list, optional): Parameters to safely pass to the query.

    Returns:
        list of dict: A list of rows as dictionaries.
    """
    try:
        with connection.cursor() as cursor:
            cursor.execute(sql_query, params or [])
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

        return [dict(zip(columns, row)) for row in rows]

    except Exception as e:
        raise Exception(f"Error executing query: {e}")

# Helper: Insert data into table
def insert_data(table_name, data):
    """
    Helper to insert data into a table.

    Args:
        table_name (str): The name of the table where data will be inserted.
        data (dict): A dictionary containing the column names as keys and the values to be inserted as values.

    Returns:
        int or None: The inserted row's ID if successful, None if failed.
    """
    try:
        columns = ', '.join(data.keys())
        placeholders = ', '.join(['%s'] * len(data))
        sql_query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"

        with transaction.atomic():
            with connection.cursor() as cursor:
                cursor.execute(sql_query, list(data.values()))
        return True # Return True if the insert was successful
                # return cursor.lastrowid  # Return inserted ID
    except Exception as e:
        raise Exception(f"Error in insert: {e}")

# Helper: Read data from table
def get_data(table_name, filters=None, search=None, search_columns=None, columns='*', limit=None, offset=None, order_by=None):
    """
    Helper to read data from a table with support for various operators like !=, IN, NOT IN, and NULL handling.
    """
    try:
        if isinstance(columns, list):
            columns = ', '.join(columns)

        sql_query = f"SELECT {columns} FROM {table_name}"

        # Adding filters (WHERE clause)
        conditions = []
        values = []

        if filters:
            for key, value in filters.items():
                if "__ne" in key:  # Not equal operator
                    column = key.replace("__ne", "")
                    if value is None:
                        conditions.append(f"{column} IS NOT NULL")
                    else:
                        conditions.append(f"{column} <> %s")
                        values.append(value)
                elif "__in" in key:  # IN operator
                    column = key.replace("__in", "")
                    if isinstance(value, (list, tuple)):
                        conditions.append(f"{column} IN ({', '.join(['%s'] * len(value))})")
                        values.extend(value)
                    else:
                        conditions.append(f"{column} = %s")
                        values.append(value)
                elif "__not_in" in key:  # NOT IN operator
                    column = key.replace("__not_in", "")
                    if isinstance(value, (list, tuple)):
                        conditions.append(f"{column} NOT IN ({', '.join(['%s'] * len(value))})")
                        values.extend(value)
                    else:
                        conditions.append(f"{column} <> %s")
                        values.append(value)
                else:
                    if value is None:
                        conditions.append(f"{key} IS NULL")
                    else:
                        conditions.append(f"{key} = %s")
                        values.append(value)

        # Adding search condition (LIKE) for the given columns
        if search and search_columns:
            search_conditions = [f"{col}::text ILIKE %s" for col in search_columns]
            conditions.append(f"({' OR '.join(search_conditions)})")
            values += [f"%{search}%"] * len(search_columns)

        # Combine all conditions
        if conditions:
            where_clause = ' WHERE ' + ' AND '.join(conditions)
            sql_query += where_clause

        # Adding limit and offset
        if limit:
            sql_query += f" LIMIT {limit}"

        if offset:
            sql_query += f" OFFSET {offset}"

        # Adding ORDER BY clause
        if order_by:
            if isinstance(order_by, list):
                order_by_clause = ', '.join(order_by)
            else:
                order_by_clause = order_by
            sql_query += f" ORDER BY {order_by_clause}"

        # Executing query
        with connection.cursor() as cursor:
            cursor.execute(sql_query, values)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

        return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        raise Exception(f"Error in read: {e}")

# Helper: Update data in a table
def update_data(table_name, data, filters):
    """
    Helper to update data in a table.

    Args:
        table_name (str): The name of the table where data will be updated.
        data (dict): The data to update, with column names as keys.
        filters (dict): Filters for the WHERE clause to specify which rows to update.

    Returns:
        bool: True if update successful, False otherwise.
    """
    try:
        set_clause = ', '.join([f"{key}=%s" for key in data.keys()])
        where_clause = ' AND '.join([f"{key}=%s" for key in filters.keys()])
        sql_query = f"UPDATE {table_name} SET {set_clause} WHERE {where_clause}"

        with transaction.atomic():
            with connection.cursor() as cursor:
                cursor.execute(sql_query, list(data.values()) + list(filters.values()))
        
        return True
    except Exception as e:
        raise Exception(f"Error in update: {e}")

# Helper: Delete data from table
def delete_data(table_name, filters):
    """
    Helper to delete data from a table.

    Args:
        table_name (str): The name of the table to delete from.
        filters (dict): Filters for the WHERE clause to specify which rows to delete.

    Returns:
        bool: True if deletion successful, False otherwise.
    """
    try:
        where_clause = ' AND '.join([f"{key}=%s" for key in filters.keys()])
        sql_query = f"DELETE FROM {table_name} WHERE {where_clause}"

        with transaction.atomic():
            with connection.cursor() as cursor:
                cursor.execute(sql_query, list(filters.values()))
        
        return True
    except Exception as e:
        raise Exception(f"Error in delete: {e}")

# Helper: Insert data into table and get the inserted row's ID
def insert_get_id_data(table_name, data, column_id):
    """
    Helper to insert data into a table and return the inserted row's ID.

    Args:
        table_name (str): The name of the table where data will be inserted.
        data (dict): A dictionary containing the column names as keys and the values to be inserted as values.

    Returns:
        int or None: The inserted row's ID if successful, None if failed.
    """
    try:
        columns = ', '.join(data.keys())
        placeholders = ', '.join(['%s'] * len(data))
        sql_query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"

        with transaction.atomic():
            with connection.cursor() as cursor:
                cursor.execute(sql_query, list(data.values()))
                # Retrieve the last inserted ID using `LAST_INSERT_ID()`
                cursor.execute(f"SELECT currval(pg_get_serial_sequence('{table_name}', '{column_id}'))")
                inserted_id = cursor.fetchone()[0]  # Get the inserted ID
        
        return inserted_id  # Return inserted ID
    except Exception as e:
        raise Exception(f"Error in insert_get_id: {e}")

# Helper: Get the first row of data from a table
def first_data(table_name, filters=None, columns='*', order_by=None):
    """
    Helper to get the first row of data from a table.

    Args:
        table_name (str): The name of the table to read from.
        filters (dict, optional): Filters for the WHERE clause.
        columns (str or list, optional): Columns to select, '*' by default.
        order_by (str or list, optional): Column or columns to order the results by.
    
    Returns:
        dict or None: A dictionary representing the first row, or None if no data found.
    """
    try:
        if isinstance(columns, list):
            columns = ', '.join(columns)

        sql_query = f"SELECT {columns} FROM {table_name}"

        # Adding filters (WHERE clause)
        where_clause = ''
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause

        # Adding order by clause
        if order_by:
            if isinstance(order_by, list):
                order_by = ', '.join(order_by)
            sql_query += f" ORDER BY {order_by}"

        sql_query += " LIMIT 1"  # Limit to the first row

        with connection.cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            columns = [col[0] for col in cursor.description]
            row = cursor.fetchone()  # Get the first row

        return dict(zip(columns, row)) if row else None
    except Exception as e:
        raise Exception(f"Error in get_first_data: {e}")

# Helper: Get the count of rows in a table
def count_data(table_name, filters=None):
    """
    Helper to get the count of rows in a table.

    Args:
        table_name (str): The name of the table to count rows from.
        filters (dict, optional): Filters for the WHERE clause.

    Returns:
        int: The count of rows that match the filters, or 0 if no data found.
    """
    try:
        sql_query = f"SELECT COUNT(*) FROM {table_name}"

        # Adding filters (WHERE clause)
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause

        with connection.cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            count = cursor.fetchone()[0]  # Get the count
        
        return count
    except Exception as e:
        raise Exception(f"Error in get_data_count: {e}")

# Helper: Pluck data from a specific column in a table
def pluck_data(table_name, column_name, filters=None):
    """
    Helper to pluck values from a specific column in a table.

    Args:
        table_name (str): The name of the table to read from.
        column_name (str): The column to pluck data from.
        filters (dict, optional): Filters for the WHERE clause.

    Returns:
        list: A list of values from the specified column.
    """
    try:
        sql_query = f"SELECT {column_name} FROM {table_name}"

        # Adding filters (WHERE clause)
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause

        with connection.cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            rows = cursor.fetchall()  # Get all matching rows

        # Extract the specified column values into a list
        return [row[0] for row in rows]
    except Exception as e:
        raise Exception(f"Error in pluck_data: {e}")

# Helper: Get distinct values from a specific column in a table
def distinct_data(table_name, column_name, filters=None):
    """
    Helper to get distinct values from a specific column in a table.

    Args:
        table_name (str): The name of the table to read from.
        column_name (str): The column to get distinct values from.
        filters (dict, optional): Filters for the WHERE clause.

    Returns:
        list: A list of distinct values from the specified column.
    """
    try:
        sql_query = f"SELECT DISTINCT {column_name} FROM {table_name}"

        # Adding filters (WHERE clause)
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause

        with connection.cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            rows = cursor.fetchall()  # Get all matching rows

        # Extract the specified column values into a list
        return [row[0] for row in rows]
    except Exception as e:
        raise Exception(f"Error in distinct_data: {e}")

# Helper: Get ordered data from a table
def order_by_data(table_name, order_column, ascending=True, filters=None, limit=None, offset=None):
    """
    Helper to get ordered data from a table.

    Args:
        table_name (str): The name of the table to read from.
        order_column (str): The column to order by.
        ascending (bool): True for ascending order, False for descending. Default is True.
        filters (dict, optional): Filters for the WHERE clause.
        limit (int, optional): Limit for the number of records.
        offset (int, optional): Offset for records.

    Returns:
        list of dict: A list of dictionaries representing the ordered rows.
    """
    try:
        sql_query = f"SELECT * FROM {table_name}"

        # Adding filters (WHERE clause)
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause

        # Adding ORDER BY clause
        order_direction = 'ASC' if ascending else 'DESC'
        sql_query += f" ORDER BY {order_column} {order_direction}"

        # Adding LIMIT and OFFSET
        if limit:
            sql_query += f" LIMIT {limit}"
        if offset:
            sql_query += f" OFFSET {offset}"

        with connection.cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()

        return [dict(zip(columns, row)) for row in rows]
    except Exception as e:
        raise Exception(f"Error in order_by_data: {e}")

# Helper: Check if data exists in a table
def exists_data(table_name, filters, id_column='id', exclude_id=None):
    """
    Helper to check if data exists in a table, with an option to exclude a specific ID.

    Args:
        table_name (str): The name of the table to check.
        filters (dict): Filters for the WHERE clause to specify which rows to check.
        id_column (str): The name of the ID column to check against.
        exclude_id (int, optional): The ID to exclude from the check.

    Returns:
        bool: True if at least one row matches the filters, False otherwise.
    """
    try:
        # Build the WHERE clause
        where_clause = ' AND '.join([f"{key}=%s" for key in filters.keys()])

        # If exclude_id is provided, add the condition to exclude it
        params = list(filters.values())
        if exclude_id is not None:
            where_clause += f' AND {id_column} != %s'
            params.append(exclude_id)  # Append exclude_id to params

        sql_query = f"SELECT EXISTS(SELECT 1 FROM {table_name} WHERE {where_clause})"

        with connection.cursor() as cursor:
            cursor.execute(sql_query, params)  # Pass all parameters
            exists = cursor.fetchone()[0]  # Fetch the first result (boolean)

        return exists
    except Exception as e:
        raise Exception(f"Error in exists_data: {e}")
    
def fetch_records_with_conditions(table_name, null_column=None, not_null_column=None, additional_filters=None):
    """
    Fetch records from a specified table based on null or not null conditions and additional filters.

    :param table_name: The name of the table to query.
    :param null_column: The name of the column to check for NULL (use for IS NULL).
    :param not_null_column: The name of the column to check for NOT NULL.
    :param additional_filters: A dictionary of additional conditions (column: value).
    :return: A list of records with the specified column(s) conditions met.
    """
    try:
        with transaction.atomic():
            with connection.cursor() as cursor:
                # Start building the SQL query
                query = f"SELECT * FROM {table_name} WHERE 1=1"  # Start with a valid base

                params = []

                # Add IS NULL condition if specified
                if null_column:
                    query += f" AND {null_column} IS NULL"

                # Add IS NOT NULL condition if specified
                if not_null_column:
                    query += f" AND {not_null_column} IS NOT NULL"

                # Add additional filters if provided
                if additional_filters:
                    conditions = []
                    for column, value in additional_filters.items():
                        conditions.append(f"{column} = %s")
                        params.append(value)
                    # Join the conditions with 'AND'
                    query += " AND " + " AND ".join(conditions)

                # Execute the query
                cursor.execute(query, params)

                # Fetch all records and convert to a list of dictionaries
                rows = cursor.fetchall()
                columns = [col[0] for col in cursor.description]  # Get column names
                result = [dict(zip(columns, row)) for row in rows]

                return result

    except Exception as e:
        raise Exception(f"Error fetching records: {str(e)}")  # Return an empty list on error
    
# Helper: Get a single value from a table
def get_value(table_name, column_name, filters=None, type=None):
    """
    Helper to get a single value from a specified column in a table.

    Args:
        table_name (str): The name of the table to read from.
        column_name (str): The column from which to retrieve the value.
        filters (dict, optional): Filters for the WHERE clause.
        type (str, optional): Type of the value, e.g., "UUID" for UUID-specific error handling.
    
    Returns:
        The value from the specified column, or raises an error if UUID is not found and type is "UUID".
    """
    try:
        sql_query = f"SELECT {column_name} FROM {table_name}"

        # Adding filters (WHERE clause)
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause

        sql_query += " LIMIT 1"  # Limit to the first row

        with connection.cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            row = cursor.fetchone()  # Get the first row

        # If type is UUID and no row is found, raise specific UUID not found error
        if type == "UUID" and row is None:
            raise Exception("UUID tidak ditemukan")

        return row[0] if row else None  # Return the value of the specified column
    except Exception as e:
        if type != "UUID":
            raise Exception(f"Error in get_value: {e}")
        else:
            raise Exception("Error in get_value: UUID tidak ditemukan")  # Specific message for UUID not found

# Helper: Save an uploaded file to a specific directory
def save_uploaded_file(file, upload_dir='uploads'):
    """
    Save an uploaded file to a specific directory.

    Args:
        file: The file object to save.
        upload_dir (str): The directory to save the uploaded files.

    Returns:
        str: The file path of the saved file.
    """
    # Buat direktori jika belum ada
    upload_path = os.path.join(settings.MEDIA_ROOT, upload_dir)
    os.makedirs(upload_path, exist_ok=True)

    # Ambil nama file asli
    original_name = os.path.splitext(file.name)[0]
    extension = os.path.splitext(file.name)[1]
    
    # Buat timestamp dalam format yang diinginkan
    timestamp = datetime.now().strftime("%d%m%y_%H%M%S_%f")
    
    # Buat nama file baru
    new_file_name = f"{original_name}_{timestamp}{extension}"
    file_path = os.path.join(upload_path, new_file_name)

    # Cek jika file sudah ada, jika ada tambahkan angka
    base, ext = os.path.splitext(new_file_name)
    counter = 1
    while os.path.exists(file_path):
        new_file_name = f"{base}_{counter}{ext}"
        file_path = os.path.join(upload_path, new_file_name)
        counter += 1

    # Simpan file
    with open(file_path, 'wb+') as destination:
        for chunk in file.chunks():
            destination.write(chunk)

    return file_path

# Helper: sum data from a column in a table
def sum_data(table_name, column_name, filters=None):
    """
    Helper to sum data from a column in a table.

    Args:
        table_name (str): The name of the table.
        column_name (str): The name of the column to sum.
        filters (dict, optional): Filters for the WHERE clause.

    Returns:
        float: The sum of the column data that matches the filters, or 0 if no data found.
    """
    try:
        sql_query = f"SELECT SUM({column_name}) FROM {table_name}"

        # Adding filters (WHERE clause)
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause

        with connection.cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            result = cursor.fetchone()[0]  # Get the sum result
            
        return result if result is not None else 0  # Return 0 if result is None (no rows)
    
    except Exception as e:
        raise Exception(f"Error in sum_data: {e}")

# Helper: Get the last row of data from a table
def last_data(table_name, filters=None, order_by_column='id', columns='*'):
    """
    Helper to get the last row of data from a table.

    Args:
        table_name (str): The name of the table to read from.
        filters (dict, optional): Filters for the WHERE clause.
        order_by_column (str, optional): The column to order the results by (default is 'id').
        columns (str or list, optional): Columns to select, '*' by default.

    Returns:
        dict or None: A dictionary representing the last row, or None if no data found.
    """
    try:
        if isinstance(columns, list):
            columns = ', '.join(columns)

        sql_query = f"SELECT {columns} FROM {table_name}"

        # Adding filters (WHERE clause)
        where_clause = ''
        if filters:
            where_clause = ' WHERE ' + ' AND '.join([f"{key}=%s" for key in filters.keys()])
            sql_query += where_clause

        # Order by the specified column in descending order
        sql_query += f" ORDER BY {order_by_column} DESC LIMIT 1"

        with connection.cursor() as cursor:
            cursor.execute(sql_query, list(filters.values()) if filters else [])
            columns = [col[0] for col in cursor.description]
            row = cursor.fetchone()  # Get the last row

        return dict(zip(columns, row)) if row else None
    except Exception as e:
        raise Exception(f"Error in last_data: {e}")

# Helper: Get the value user_id from uuid jwt
def jwt_uuid_conveter(uuid):
    user_id = get_value(
        table_name='sso.users',
        filters={
            'user_uuid': uuid
        },
        column_name="user_id"
    )
    
    if not user_id:
        return None
    return user_id

# Helper: Validate Request json from FE
def validate_request(data, rules):
    errors = {}

    for field, validations in rules.items():
        field_validations = validations.split('|')

        for validation in field_validations:
            # Required rule
            if validation == 'required' and field not in data:
                errors[field] = f"{field} is required."
                continue
            
            # Skip further checks if field is not present and it's not required
            if field not in data:
                continue
            
            value = data[field]

            # String rule
            if validation == 'string' and not isinstance(value, str):
                errors[field] = f"{field} must be a string."
            
            # Numeric rule
            if validation == 'numeric' and not isinstance(value, (int, float)):
                errors[field] = f"{field} must be numeric."
            
            # Boolean rule
            if validation == 'boolean' and not isinstance(value, bool):
                errors[field] = f"{field} must be true or false."
            
            # Email rule
            if validation == 'email':
                email_regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
                if not re.match(email_regex, value):
                    errors[field] = f"{field} must be a valid email."

            # Date rule
            if validation == 'date':
                try:
                    datetime.strptime(value, '%Y-%m-%d')
                except ValueError:
                    errors[field] = f"{field} must be a valid date (YYYY-MM-DD)."
            
            # Min length for strings or min value for numbers
            if validation.startswith('min:'):
                min_value = int(validation.split(':')[1])
                if isinstance(value, (int, float)) and value < min_value:
                    errors[field] = f"{field} must be at least {min_value}."
                if isinstance(value, str) and len(value) < min_value:
                    errors[field] = f"{field} must be at least {min_value} characters long."

            # Max length for strings or max value for numbers
            if validation.startswith('max:'):
                max_value = int(validation.split(':')[1])
                if isinstance(value, (int, float)) and value > max_value:
                    errors[field] = f"{field} must be no more than {max_value}."
                if isinstance(value, str) and len(value) > max_value:
                    errors[field] = f"{field} must be no more than {max_value} characters long."
            
            # In rule (check if value is in a list of allowed values)
            if validation.startswith('in:'):
                allowed_values = validation.split(':')[1].split(',')
                if value not in allowed_values:
                    errors[field] = f"{field} must be one of {', '.join(allowed_values)}."

            # Array rule
            if validation == 'array' and not isinstance(value, list):
                errors[field] = f"{field} must be an array."
            
            # Integer rule
            if validation == 'integer' and not isinstance(value, int):
                errors[field] = f"{field} must be an integer."

            # Float rule
            if validation == 'float' and not isinstance(value, float):
                errors[field] = f"{field} must be a floating-point number."
            
            # Digits rule (e.g. digits:5 means must be a 5-digit number)
            if validation.startswith('digits:'):
                num_digits = int(validation.split(':')[1])
                if not (isinstance(value, int) and len(str(value)) == num_digits):
                    errors[field] = f"{field} must be a {num_digits}-digit number."
            
            # URL rule
            if validation == 'url':
                url_regex = r'^(https?|ftp)://[^\s/$.?#].[^\s]*$'
                if not re.match(url_regex, value):
                    errors[field] = f"{field} must be a valid URL."

            # Starts with rule
            if validation.startswith('starts_with:'):
                prefix = validation.split(':')[1]
                if not value.startswith(prefix):
                    errors[field] = f"{field} must start with {prefix}."
            
            # Ends with rule
            if validation.startswith('ends_with:'):
                suffix = validation.split(':')[1]
                if not value.endswith(suffix):
                    errors[field] = f"{field} must end with {suffix}."

            # Unique rule (simulate by checking a list of existing values)
            if validation == 'unique':
                # Normally, you'd check a database for uniqueness
                # Simulate by checking against a list of existing values
                existing_values = ['example1', 'example2']  # Replace with actual database check
                if value in existing_values:
                    errors[field] = f"{field} must be unique."
    
    # Return errors if any are found
    if errors:
        return errors
    
    # No errors, validation passed
    return None

def validate_method(request, required_method):
    if request.method != required_method:
        raise ValueError(f"Method not allowed. Please use {required_method} method.")
    

def log_exception(request, exception):
    try:
        # Parse traceback to get the file and line where the error occurred
        tb = traceback.extract_tb(exception.__traceback__)
        if tb:
            file_location = f"{tb[0].filename} line {tb[0].lineno}"
            error_message = f"{tb[-1].filename} line {tb[-1].lineno}: {tb[-1].line} - {str(exception)}"
        else:
            file_location = "Unknown file"
            error_message = f"{tb[-1].filename} line {tb[-1].lineno}: {tb[-1].line} - {str(exception)}"

        # Convert QueryDict to a JSON string
        query_params = json.dumps(request.GET.dict())

        # Insert log entry
        insert_data(
            table_name="log.tbl_log",
            data={
                "method": request.method,
                "path": request.path,
                "query_params": query_params,
                "client_ip": request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR')),
                "user_agent": request.META.get('HTTP_USER_AGENT'),
                "employee_uuid": getattr(request, 'jwt_uuid', None),
                "employee_no": getattr(request, 'jwt_badge_no', None),
                "employee_name": getattr(request, 'jwt_fullname', None),
                "file_location": file_location,
                "message_error": error_message,
                "created_at": datetime.now()
            }
        )
    except Exception as e:
        print(f"Failed to log exception: {str(e)}")


def generate_custom_excel(data, headers, title, file_name, dir):
    """
    Generate an Excel file with customizable content.

    Parameters:
    - data (list of dict): The data to populate the Excel file, where each dict represents a row.
    - headers (list of str): Column headers for the Excel sheet.
    - title (str): Title for the report.
    - file_name (str): The name of the file to save.
    - dir (str): The directory to save the file in.
    """
    
    # Create a new workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Whitelist Metal Detector"

    # Title Row
    ws['A1'] = title
    ws['A1'].font = Font(size=20, bold=True, color="000000")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Subtitle Row with timestamp
    timestamp = datetime.now().strftime('%d %b %Y %H:%M:%S')
    ws['A2'] = f"Generated at : {timestamp}"
    ws['A2'].font = Font(size=11, italic=True, color="555555")
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    # Header Row (starts from W4)
    header_row_index = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_index, column=col_num, value=header)
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

    # Populate Data Rows
    for row_index, row_data in enumerate(data, start=header_row_index + 1):
        for col_index, header in enumerate(headers, start=1):
            cell_value = row_data.get(header)

            # Check if the value is a date (datetime or date object)
            if isinstance(cell_value, (datetime, date)):
                # Format the date to the desired format (e.g., 'dd-mm-yyyy')
                cell_value = cell_value.strftime('%d-%m-%Y')

            cell = ws.cell(row=row_index, column=col_index, value=cell_value)
            cell.alignment = Alignment(horizontal="left" if header in ["Employee Name", "Description"] else "center", vertical="center")
            cell.border = Border(
                top=Side(border_style="thin", color="000000"),
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=2, max_col=2):  # (Column B)
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=3, max_col=3):  # (Column C)
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center") 
    
    # Adjust column widths
    column_widths = {
        "A": 10, "B": 40, "C": 60, "D": 30, "E": 15
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    export_dir = os.path.join(os.path.dirname(__file__), f'{dir}')
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)

    # Output file path
    output_path = os.path.join(export_dir, f'{file_name}.xlsx')

    # If file exists, add timestamp to filename
    if os.path.exists(output_path):
        timestamp = datetime.now().strftime('%d_%b_%Y_%H_%M_%S')
        output_path = os.path.join(export_dir, f'{file_name.split(".")[0]}_{timestamp}.xlsx')

    # Save the workbook
    wb.save(output_path)
    print(f"Excel file saved at {output_path}")
    
def generate_excel_from_template(data, url=None, output_file_name="dashboard_report.xlsx"):
    """
    Generate an Excel file from data, including dynamic filter and report title.
    :param data: Data to be written in the Excel report (list of sheets, each containing 'sheet_name', 'headers', 'data').
    :param url: URL string to extract filter parameters, default is None.
    :param output_file_name: Name of the output Excel file, default is "dashboard_report.xlsx".
    :return: The path to the generated Excel file.
    """
    # Function to extract parameters from URL
    def extract_url_params(url):
        if not url:
            return "", "", "", ""  # Default empty values if URL is None

        parsed_url = urlparse(url)
        query_params = parse_qs(parsed_url.query)

        # Extract individual query parameters
        year = query_params.get('year', [''])[0]
        month = query_params.get('month', [''])[0]
        department = query_params.get('department', [''])[0]
        floor = query_params.get('floor', [''])[0]

        return year, month, department, floor

    # Extract parameters from URL (if provided)
    year, month, department, floor = extract_url_params(url)

    # Set up output folder to be the user's Downloads folder
    home_dir = os.path.expanduser("~")
    downloads_folder = os.path.join(home_dir, "Downloads")

    # Ensure the Downloads directory exists
    if not os.path.exists(downloads_folder):
        os.makedirs(downloads_folder)

    output_file = os.path.join(downloads_folder, output_file_name)

    # Check if file already exists and add a number if needed
    base_name, ext = os.path.splitext(output_file_name)
    counter = 1
    while os.path.exists(output_file):
        output_file = os.path.join(downloads_folder, f"{base_name}({counter}){ext}")
        counter += 1

    try:
        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            for sheet in data:
                sheet_name = sheet['sheet_name']
                headers = sheet['headers']
                data_rows = sheet['data']

                # Reformat the data to match the required format
                formatted_data = []
                for row in data_rows:
                    formatted_row = [row['label']]  # The first column will contain values like "Owl", "Nowl", "Total"
                    for month in headers:
                        formatted_row.append(row.get(month, 0))  # Default to 0 if no value exists
                    formatted_data.append(formatted_row)

                # Create the DataFrame with the first column as 'label' and the months as headers
                df = pd.DataFrame(formatted_data, columns=['label'] + headers)

                # Write the data to the Excel file without the header (since we're writing it manually)
                df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=4, header=False)

                # Access the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Header format (for the months)
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'align': 'center',
                    'valign': 'vcenter',
                    'bg_color': '#FFFF00',
                    'border': 1
                })

                # Title format (for the sheet title)
                title_format = workbook.add_format({
                    'bold': True,
                    'font_size': 16,
                    'align': 'center',
                    'valign': 'vcenter',
                    'font_color': '#000000'
                })

                # Filter label format
                label_format = workbook.add_format({
                    'bold': True,
                    'font_size': 12,
                    'align': 'left',
                    'valign': 'vcenter',
                })

                # Dynamically calculate the column range based on headers length
                last_col = len(headers)
                title_range = f"A1:{chr(65 + last_col)}1"  # Adjust for the column length dynamically
                worksheet.merge_range(title_range, f"{sheet_name}", title_format)

                # Create dynamic filter string, excluding empty filters
                filters = []
                if year:
                    filters.append(f"{year}")
                if month:
                    filters.append(f"{month}")
                if department:
                    filters.append(f"{department}")
                if floor:
                    filters.append(f"{floor}")

                # Join the filters with commas if any are present
                filter_by_text = f"Filter by: {', '.join(filters)}" if filters else "Filter by: N/A"
                generate_on_text = f"Generate on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

                # Write the filter and generation timestamp
                worksheet.write('A2', filter_by_text, label_format)
                worksheet.write('A3', generate_on_text, label_format)

                # Write the headers in row 4 (just below the filter)
                for col_num, header in enumerate(headers, start=1):
                    worksheet.write(3, col_num, header, header_format)

                # Auto-adjust column widths based on the content
                for i, column in enumerate(df.columns):
                    max_width = max(df[column].astype(str).map(len).max(), len(column)) + 2
                    worksheet.set_column(i, i, max_width)

                # Apply border around the entire data range (excluding title and filter rows)
                last_row = len(formatted_data) + 4  # Data starts from row 4
                last_col = len(headers) + 1  # Including 'label' column

                # Set border around the data cells (A4:Z[last_row])
                border_format = workbook.add_format({'border': 1})

                # Apply border only to the data range (A4:Z[last_row])
                worksheet.conditional_format(f"A4:{chr(64 + last_col)}{last_row}", {
                    'type': 'no_blanks',
                    'format': border_format
                })

                # Apply specific color to cells in column A for rows 5-7 without overwriting data
                row_color_format = workbook.add_format({
                    'bg_color': '#dddedd'
                })
                worksheet.conditional_format("A5:A7", {
                    'type': 'no_blanks',
                    'format': row_color_format
                })

        return output_file

    except Exception as e:
        raise Exception(f"Error saving Excel file: {e}")